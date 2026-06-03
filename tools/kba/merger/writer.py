"""
Modulo writer per il tool kba.merger.

Fornisce funzioni per scrivere i report del gap check in formato TXT e Excel.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from loguru import logger

from tools.kba.merger.config import OUTPUT_DIR, GAP_STATUS_COLORS


def write_gap_txt(gap_rows: list[dict[str, Any]], today: str) -> Path:
    """
    Scrive un report TXT del gap check.

    Args:
        gap_rows: Lista dei risultati del gap check.
        today: Data ISO per il nome file.

    Returns:
        Path del file TXT creato.
    """
    output_path = OUTPUT_DIR / f"gap_report_{today}.txt"
    lines = [
        "Gap Check Report",
        f"Generated: {today}",
        "",
        "KBA Number\tOccurrences\tStatus\tReferenced By",
        "-" * 80,
    ]

    for row in gap_rows:
        ref_by = ", ".join(row["referenced_by"]) if row["referenced_by"] else ""
        line = f"{row['kba_number']}\t{row['occurrences']}\t{row['stato']}\t{ref_by}"
        lines.append(line)

    content = "\n".join(lines)
    output_path.write_text(content, encoding="utf-8")
    logger.info(f"Report TXT scritto: {output_path}")
    return output_path


def write_gap_excel(gap_rows: list[dict[str, Any]], today: str) -> Path:
    """
    Scrive un report Excel del gap check.

    Args:
        gap_rows: Lista dei risultati del gap check.
        today: Data ISO per il nome file.

    Returns:
        Path del file Excel creato.
    """
    output_path = OUTPUT_DIR / f"gap_report_{today}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gap Check"

    # Header
    headers = [
        "KBA Number",
        "Published",
        "Category",
        "Disposition Status",
        "Title",
        "Occurrences",
        "Status",
        "Referenced By",
    ]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Dati
    for row_num, row in enumerate(gap_rows, 2):
        ws.cell(row=row_num, column=1, value=row["kba_number"])
        ws.cell(row=row_num, column=2, value=row["published"])
        ws.cell(row=row_num, column=3, value=row["category"])
        ws.cell(row=row_num, column=4, value=row["disposition_status"])
        ws.cell(row=row_num, column=5, value=row["title"])
        ws.cell(row=row_num, column=6, value=row["occurrences"])
        ws.cell(row=row_num, column=7, value=row["stato"])
        ws.cell(
            row=row_num,
            column=8,
            value=", ".join(row["referenced_by"]) if row["referenced_by"] else "",
        )

        # Colora la colonna Status
        status_cell = ws.cell(row=row_num, column=7)
        if row["stato"] in GAP_STATUS_COLORS:
            status_cell.fill = PatternFill(
                start_color=GAP_STATUS_COLORS[row["stato"]],
                end_color=GAP_STATUS_COLORS[row["stato"]],
                fill_type="solid",
            )

    # Larghezze colonne
    for col_num, width in enumerate([14, 12, 12, 14, 50, 10, 12, 30], 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    wb.save(output_path)
    logger.info(f"Report Excel scritto: {output_path}")
    return output_path


def write_merge_excel(enriched_rows: list[dict[str, Any]], output_path: Path | None = None) -> Path:
    """
    Scrive l'Excel merged e enriched.

    Args:
        enriched_rows: Lista dei dati enriched.
        output_path: Path di output opzionale.

    Returns:
        Path del file Excel creato.
    """
    if output_path is None:
        from tools.kba.merger.config import OUTPUT_DIR
        from datetime import datetime

        today = datetime.now().strftime("%y%m%d")
        output_path = OUTPUT_DIR / f"merged_enriched_{today}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged Enriched"

    if not enriched_rows:
        wb.save(output_path)
        logger.info(f"Excel vuoto scritto: {output_path}")
        return output_path

    # Header dalle chiavi del primo row
    headers = list(enriched_rows[0].keys())
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Dati
    for row_num, row in enumerate(enriched_rows, 2):
        for col_num, header in enumerate(headers, 1):
            value = row.get(header, "")
            ws.cell(row=row_num, column=col_num, value=value)

    # Larghezze colonne auto
    for col_num in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        for row_num in range(1, len(enriched_rows) + 2):
            cell_value = str(ws.cell(row=row_num, column=col_num).value or "")
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)  # Max 50

    wb.save(output_path)
    logger.info(f"Excel merged scritto: {output_path}")
    return output_path
