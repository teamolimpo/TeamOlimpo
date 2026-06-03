"""
Subcommand learn: legge un Excel revisionato (con FIS Notes compilate),
costruisce/aggiorna il prontuario in lib/data/kba_prontuario/.

Struttura prontuario:
  lib/data/kba_prontuario/
    cases/         — un file .md per ogni KBA revisionata
    index.yaml     — statistiche aggregate
    rules.md       — regole apprese dai disaccordi (generato da LLM se disponibile)
"""

from __future__ import annotations

import re
import sys
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
import yaml
from loguru import logger

from tools.kba.merger.config import PROJECT_ROOT

# ---------------------------------------------------------------------------
# Costanti
# ---------------------------------------------------------------------------

PRONTUARIO_DIR: Path = PROJECT_ROOT / "lib" / "data" / "kba_prontuario"
CASES_DIR: Path = PRONTUARIO_DIR / "cases"

# Pattern tag in FIS Notes: {SIGLA keyword}
# Es: {RS done}, {FIS na}, {Mario std}
TAG_PATTERN = re.compile(r"\{(\w+)\s+(std|done|na|defer)\}", re.IGNORECASE)

TAG_TO_ACTION: dict[str, str] = {
    "std": "In-Progress",
    "done": "Action Complete",
    "na": "Not Applicable",
    "defer": "Deferred",
}


# ---------------------------------------------------------------------------
# Parsing tag FIS Notes
# ---------------------------------------------------------------------------


def parse_fis_tag(fis_notes: str) -> tuple[str, str, str]:
    """
    Estrae (author, keyword, human_action) da un testo FIS Notes.

    Cerca il pattern {SIGLA keyword} nel testo (case-insensitive).
    Se nessun tag trovato: restituisce ("", "", "") — significa conferma AI.

    Args:
        fis_notes: Testo della cella FIS Notes revisionata.

    Returns:
        Tupla (author, keyword, human_action) o ("", "", "") se nessun tag.

    Examples:
        >>> parse_fis_tag("{RS done} Patch NK-2400-0160 applicata il 15/03/2026")
        ('RS', 'done', 'Action Complete')
        >>> parse_fis_tag("nessun tag qui")
        ('', '', '')
    """
    if not fis_notes or not fis_notes.strip():
        return ("", "", "")

    match = TAG_PATTERN.search(fis_notes)
    if not match:
        return ("", "", "")

    author = match.group(1).upper()
    keyword = match.group(2).lower()
    human_action = TAG_TO_ACTION.get(keyword, "")

    if not human_action:
        logger.warning(f"Keyword tag non riconosciuta: {keyword!r}")
        return ("", "", "")

    return (author, keyword, human_action)


# ---------------------------------------------------------------------------
# Lettura Excel revisionato
# ---------------------------------------------------------------------------


def _read_merged_excel(xlsx_path: Path) -> list[dict[str, Any]]:
    """
    Legge il foglio KBA_Merged dall'Excel revisionato.

    Cerca le colonne per nome header (non per indice fisso).

    Args:
        xlsx_path: Path al file .xlsx revisionato.

    Returns:
        Lista di dizionari {header: valore}.

    Raises:
        ValueError: Se il foglio KBA_Merged non esiste o gli header mancano.
        OSError: Se il file non e' leggibile.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Cerca foglio KBA_Merged (case-insensitive)
    ws = None
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == "kba_merged":
            ws = wb[sheet_name]
            break

    if ws is None:
        # Fallback: foglio attivo
        ws = wb.active
        if ws is None:
            raise ValueError(f"Nessun foglio trovato in: {xlsx_path.name}")
        logger.warning(f"Foglio 'KBA_Merged' non trovato, uso foglio attivo: {ws.title}")

    rows_iter = ws.iter_rows(values_only=True)

    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        raise ValueError(f"File vuoto: {xlsx_path.name}")

    headers = [str(h).strip() if h is not None else "" for h in raw_headers]

    rows: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        if all(v is None or str(v).strip() == "" for v in raw_row):
            continue
        row_dict: dict[str, Any] = {}
        for col_idx, header in enumerate(headers):
            value = raw_row[col_idx] if col_idx < len(raw_row) else None
            row_dict[header] = value
        rows.append(row_dict)

    wb.close()
    logger.debug(f"Lette {len(rows)} righe dal foglio revisionato")
    return rows


# ---------------------------------------------------------------------------
# Scrittura case file
# ---------------------------------------------------------------------------


def _case_slug(kba_id: str, site: str) -> str:
    """
    Genera uno slug univoco per il file case.

    Args:
        kba_id: Es. "NK-2400-0150"
        site: Es. "Montecchio"

    Returns:
        Stringa slug lowercase, es. "nk-2400-0150_montecchio"
    """
    site_part = re.sub(r"[^\w]", "_", site.lower()) if site else "unknown"
    return f"{kba_id.lower()}_{site_part}"


def _write_case_file(
    slug: str,
    kba_id: str,
    risk_level: str,
    ai_action: str,
    ai_confidence: str,
    ai_rationale: str,
    human_action: str,
    human_author: str,
    fis_notes: str,
    reviewed_at: str,
) -> Path:
    """
    Scrive o sovrascrive il file case in CASES_DIR/<slug>.md.

    Args:
        slug: Identificatore univoco del caso.
        kba_id: Numero KBA.
        risk_level: Livello di rischio.
        ai_action: Azione raccomandata dall'AI.
        ai_confidence: Confidenza AI.
        ai_rationale: Motivazione AI.
        human_action: Azione decisa dall'operatore umano.
        human_author: Sigla autore (dal tag FIS).
        fis_notes: Testo completo FIS Notes.
        reviewed_at: Data revisione (ISO format).

    Returns:
        Path al file scritto.
    """
    CASES_DIR.mkdir(parents=True, exist_ok=True)

    agreement = ai_action == human_action

    # Frontmatter YAML
    fm: dict[str, Any] = {
        "kba_id": kba_id,
        "risk_level": risk_level,
        "ai_action": ai_action,
        "ai_confidence": ai_confidence,
        "human_action": human_action,
        "human_author": human_author,
        "agreement": agreement,
        "fis_notes": fis_notes,
        "reviewed_at": reviewed_at,
    }

    fm_str = yaml.dump(fm, allow_unicode=True, default_flow_style=False, sort_keys=False)

    content = (
        f"---\n{fm_str}---\n\n"
        f"## Rationale AI\n{ai_rationale or '(non disponibile)'}\n\n"
        f"## Decisione umana\n{fis_notes or '(nessuna nota)'}\n"
    )

    case_path = CASES_DIR / f"{slug}.md"
    case_path.write_text(content, encoding="utf-8")
    logger.debug(f"Case scritto: {case_path.name} | agreement={agreement}")
    return case_path


# ---------------------------------------------------------------------------
# Aggiornamento index.yaml
# ---------------------------------------------------------------------------


def _update_index(cases_data: list[dict[str, Any]]) -> None:
    """
    Riscrive index.yaml con le statistiche aggregate di tutti i casi.

    Args:
        cases_data: Lista di dizionari con i campi di ogni caso processato.
    """
    PRONTUARIO_DIR.mkdir(parents=True, exist_ok=True)

    total = len(cases_data)
    agreements = sum(1 for c in cases_data if c["agreement"])
    disagreements = total - agreements
    agreement_rate = round(agreements / total, 4) if total else 0.0

    authors: dict[str, dict[str, int]] = {}
    for c in cases_data:
        author = c.get("human_author") or "unknown"
        if author not in authors:
            authors[author] = {"total": 0, "agreements": 0}
        authors[author]["total"] += 1
        if c["agreement"]:
            authors[author]["agreements"] += 1

    index: dict[str, Any] = {
        "last_updated": date.today().isoformat(),
        "total_cases": total,
        "agreement_rate": agreement_rate,
        "disagreements": disagreements,
        "authors": authors,
    }

    index_path = PRONTUARIO_DIR / "index.yaml"
    index_path.write_text(
        yaml.dump(index, allow_unicode=True, default_flow_style=False, sort_keys=False),
        encoding="utf-8",
    )
    logger.debug(f"index.yaml aggiornato: {total} casi, agreement_rate={agreement_rate:.2%}")


# ---------------------------------------------------------------------------
# Rigenerazione rules.md
# ---------------------------------------------------------------------------


def _regen_rules(
    disagreements: list[dict[str, Any]],
    provider: Any,
    model: str | None,
    total_cases: int,
    agreement_rate: float,
) -> bool:
    """
    Rigenera rules.md usando il provider LLM sui casi di disaccordo.

    Args:
        disagreements: Lista di dizionari dei casi con agreement=False.
        provider: Istanza provider LLM.
        model: Override modello.
        total_cases: Numero totale di casi (per intestazione).
        agreement_rate: Tasso di accordo (per intestazione).

    Returns:
        True se rules.md e' stato riscritto, False in caso di errore.
    """
    learn_prompt_path = PROJECT_ROOT / "lib" / "Prompts" / "kba" / "estrai-regole-prontuario.md"
    if not learn_prompt_path.exists():
        logger.warning(f"Prompt learn non trovato: {learn_prompt_path}")
        return False

    # Estrae sezione ## Prompt
    content = learn_prompt_path.read_text(encoding="utf-8")
    import re as _re

    match = _re.search(r"^## Prompt\s*\n([\s\S]*?)(?=\n## |\Z)", content, _re.MULTILINE)
    if not match:
        logger.warning("Sezione '## Prompt' non trovata nel prompt learn")
        return False

    prompt_template = match.group(1).strip()

    # Costruisce il testo dei casi di disaccordo
    case_lines: list[str] = []
    for i, c in enumerate(disagreements, 1):
        case_lines.append(
            f"--- Caso {i} ---\n"
            f"KBA ID: {c['kba_id']}\n"
            f"Risk Level: {c['risk_level']}\n"
            f"AI Action: {c['ai_action']} (confidence: {c['ai_confidence']})\n"
            f"Human Action: {c['human_action']} (autore: {c['human_author']})\n"
            f"FIS Notes: {c['fis_notes']}\n"
            f"AI Rationale: {c['ai_rationale']}\n"
        )

    kba_text = "\n".join(case_lines)
    prompt = prompt_template.replace("{{kba_text}}", kba_text)

    try:
        response = provider.chat(prompt=prompt, model=model)
        rules_body = response.text.strip()
    except Exception as exc:
        logger.warning(f"Errore rigenerazione rules.md: {exc}")
        return False

    today = date.today().isoformat()
    pct = f"{agreement_rate * 100:.0f}%"
    header = (
        f"# Regole prontuario KBA\n\n"
        f"Aggiornato: {today} | Casi totali: {total_cases} | Agreement rate: {pct}\n\n"
        f"## Regole apprese\n\n"
    )

    rules_path = PRONTUARIO_DIR / "rules.md"
    rules_path.write_text(header + rules_body + "\n", encoding="utf-8")
    logger.info(f"rules.md riscritto: {rules_path}")
    return True


def _write_rules_placeholder(total_cases: int, agreement_rate: float) -> None:
    """
    Scrive un rules.md placeholder quando il provider non e' disponibile.

    Args:
        total_cases: Numero totale di casi.
        agreement_rate: Tasso di accordo.
    """
    PRONTUARIO_DIR.mkdir(parents=True, exist_ok=True)
    today = date.today().isoformat()
    pct = f"{agreement_rate * 100:.0f}%"
    content = (
        f"# Regole prontuario KBA\n\n"
        f"Aggiornato: {today} | Casi totali: {total_cases} | Agreement rate: {pct}\n\n"
        f"## Regole apprese\n\n"
        f"_(Nessun provider disponibile — esegui `learn --provider grok` per generare le regole)_\n"
    )
    rules_path = PRONTUARIO_DIR / "rules.md"
    rules_path.write_text(content, encoding="utf-8")


# ---------------------------------------------------------------------------
# Entry point principale
# ---------------------------------------------------------------------------


def run_learn(
    input_xlsx: Path,
    provider: Any = None,
    model: str | None = None,
) -> None:
    """
    Legge il foglio KBA_Merged dall'Excel revisionato.

    Per ogni riga con colonne AI Action (P) e FIS Notes (O/S):
    - Determina human_action dal tag in FIS Notes (o = ai_action se nessun tag)
    - Scrive/aggiorna cases/<slug>.md
    - Aggiorna index.yaml
    - Se ci sono disaccordi E provider disponibile: rigenera rules.md

    Output stderr:
      Casi letti, nuovi casi, gia' in prontuario, disaccordi, stato rules.md

    Args:
        input_xlsx: Path al file .xlsx revisionato.
        provider: Istanza provider LLM (opzionale).
        model: Override modello LLM (opzionale).

    Raises:
        ValueError: Se il file Excel non e' leggibile o mancano colonne attese.
        OSError: Se non e' possibile scrivere sul filesystem.
    """
    rows = _read_merged_excel(input_xlsx)
    today = date.today().isoformat()

    # Conta casi esistenti prima del run
    existing_slugs: set[str] = set()
    if CASES_DIR.exists():
        existing_slugs = {p.stem for p in CASES_DIR.glob("*.md")}

    total_read = len(rows)
    new_cases = 0
    already_in = 0
    all_cases_data: list[dict[str, Any]] = []
    disagree_cases: list[dict[str, Any]] = []

    for row in rows:
        kba_id = str(row.get("KBA Number") or "").strip()
        if not kba_id:
            continue

        site = str(row.get("Site") or "").strip()
        slug = _case_slug(kba_id, site)

        # Legge colonne AI (P, Q, R) — cerca per nome header
        ai_action = str(row.get("Recommended Action") or "").strip()
        ai_rationale = str(row.get("Rationale") or "").strip()
        ai_confidence = str(row.get("Confidence") or "").strip()

        # FIS Notes puo' essere in "FIS Notes" o "S" dopo riordino colonne
        fis_notes = str(row.get("FIS Notes") or "").strip()

        risk_level = str(row.get("Risk Level") or "").strip()

        # Determina human_action
        author, keyword, human_action = parse_fis_tag(fis_notes)

        if not human_action:
            # Nessun tag = conferma la raccomandazione AI
            human_action = ai_action
            author = ""

        if not ai_action:
            # Riga senza raccomandazione AI: salta
            logger.debug(f"Riga {kba_id} senza Recommended Action — saltata")
            continue

        case_info: dict[str, Any] = {
            "slug": slug,
            "kba_id": kba_id,
            "risk_level": risk_level,
            "ai_action": ai_action,
            "ai_confidence": ai_confidence,
            "ai_rationale": ai_rationale,
            "human_action": human_action,
            "human_author": author,
            "fis_notes": fis_notes,
            "agreement": (ai_action == human_action),
        }

        all_cases_data.append(case_info)

        if slug in existing_slugs:
            already_in += 1
        else:
            new_cases += 1

        if not case_info["agreement"]:
            disagree_cases.append(case_info)

        # Scrive il file case
        _write_case_file(
            slug=slug,
            kba_id=kba_id,
            risk_level=risk_level,
            ai_action=ai_action,
            ai_confidence=ai_confidence,
            ai_rationale=ai_rationale,
            human_action=human_action,
            human_author=author,
            fis_notes=fis_notes,
            reviewed_at=today,
        )

    # Aggiorna index
    _update_index(all_cases_data)

    total_cases = len(all_cases_data)
    agreements = sum(1 for c in all_cases_data if c["agreement"])
    agreement_rate = agreements / total_cases if total_cases else 0.0
    disagree_count = len(disagree_cases)

    # Rigenera rules.md
    rules_regenerated = False
    if disagree_cases and provider is not None:
        rules_regenerated = _regen_rules(
            disagreements=disagree_cases,
            provider=provider,
            model=model,
            total_cases=total_cases,
            agreement_rate=agreement_rate,
        )
    elif not (PRONTUARIO_DIR / "rules.md").exists():
        _write_rules_placeholder(total_cases, agreement_rate)

    # Report stderr
    disagree_pct = f"({disagree_count / total_cases * 100:.0f}%)" if total_cases else ""
    rules_status = (
        "rules.md rigenerato"
        if rules_regenerated
        else (
            "rules.md non rigenerato (provider non disponibile)"
            if disagree_cases
            else "rules.md invariato"
        )
    )

    sys.stderr.write(
        f"Casi letti:          {total_read}\n"
        f"Nuovi casi:           {new_cases}\n"
        f"Gia' in prontuario:    {already_in}\n"
        f"Disaccordi:           {disagree_count}  {disagree_pct}\n"
        f"{rules_status}\n"
        f"Prontuario: {PRONTUARIO_DIR}\n"
    )
