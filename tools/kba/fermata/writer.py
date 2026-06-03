"""
Scrive il file Excel fermata con 4 sheet: Files, Montecchio, Lonigo, Termoli.

Struttura input patch_data (da kba.reporter.patch_builder.build_patch_list):
{
  "Lonigo": {
    "workstation_ms": {"nome_file.zip": ["N1","N2"]},
    "server_ms":      {"nome_file.zip": ["S1"]},
    "firmware":       {"nome_file_CSS": {"type":"controller","nodes":["C1"]}},
  },
  "Montecchio": {...},
  "Termoli":    {...},
}
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SITE_ORDER = ["Montecchio", "Lonigo", "Termoli"]

HEADER_BG = "1F4E79"
HEADER_FG = "FFFFFF"
FILE_BG = "D6E4F7"  # azzurro chiaro — sheet Files
WS_BG = "E2EFDA"  # verde chiaro  — workstation
SRV_BG = "FFF2CC"  # giallo chiaro — server
FW_BG = "FCE4D6"  # arancio chiaro — firmware/controller


def _hfill(hex_: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_)


def _hfont(color: str = HEADER_FG, bold: bool = True) -> Font:
    return Font(bold=bold, color=color)


def _top() -> Alignment:
    return Alignment(vertical="top")


def _wrap() -> Alignment:
    return Alignment(wrap_text=True, vertical="top")


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _add_table(ws, name: str, ref: str) -> None:
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True,
        showColumnStripes=False,
        showFirstColumn=False,
        showLastColumn=False,
    )
    ws.add_table(t)


# ---------------------------------------------------------------------------
# Sheet "Files" — lista deduplicata di tutti i file da procurare
# ---------------------------------------------------------------------------


def _collect_all_files(patch_data: dict) -> list[dict]:
    """
    Aggrega tutti i file presenti nei tre siti, deduplicando.
    Ritorna lista di dict {file, tipo, siti} ordinata per tipo poi nome.
    """
    all_files: dict[str, dict] = {}
    for site, data in patch_data.items():
        for f in data.get("workstation_ms", {}):
            entry = all_files.setdefault(f, {"file": f, "tipo": "Workstation/Server", "siti": []})
            if site not in entry["siti"]:
                entry["siti"].append(site)
        for f in data.get("server_ms", {}):
            entry = all_files.setdefault(f, {"file": f, "tipo": "Server", "siti": []})
            if site not in entry["siti"]:
                entry["siti"].append(site)
        for f in data.get("firmware", {}):
            ftype = data["firmware"][f].get("type", "io")
            tipo = "Firmware Controller" if ftype == "controller" else "Firmware I/O"
            entry = all_files.setdefault(f, {"file": f, "tipo": tipo, "siti": []})
            if site not in entry["siti"]:
                entry["siti"].append(site)

    result = sorted(all_files.values(), key=lambda x: (x["tipo"], x["file"]))
    return result


def _write_files_sheet(ws, patch_data: dict) -> None:
    headers = ["File", "Tipo", "Siti"]
    widths = [60, 22, 30]

    hfill = _hfill(HEADER_BG)
    hfont = _hfont()
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = _top()

    rows = _collect_all_files(patch_data)
    for ri, r in enumerate(rows, 2):
        siti_str = ", ".join(s for s in SITE_ORDER if s in r["siti"])
        ws.cell(row=ri, column=1, value=r["file"]).alignment = _top()
        ws.cell(row=ri, column=2, value=r["tipo"]).alignment = _top()
        ws.cell(row=ri, column=3, value=siti_str).alignment = _top()
        for ci in range(1, 4):
            ws.cell(row=ri, column=ci).fill = _hfill(FILE_BG)

    _set_col_widths(ws, widths)
    ws.freeze_panes = "A2"
    if rows:
        _add_table(ws, "Files_Table", f"A1:C{len(rows) + 1}")


# ---------------------------------------------------------------------------
# Sheet per sito — file con nodi/controller
# ---------------------------------------------------------------------------


def _write_site_sheet(ws, site: str, site_data: dict) -> None:
    headers = ["File", "Tipo", "Nodi / Note"]
    widths = [60, 22, 50]

    hfill = _hfill(HEADER_BG)
    hfont = _hfont()
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = _top()

    row_idx = 2
    site_rows = []

    # Workstation
    for f, nodes in sorted(site_data.get("workstation_ms", {}).items()):
        site_rows.append(
            {
                "file": f,
                "tipo": "Workstation",
                "nodi": ", ".join(nodes) if nodes else "tutti",
                "bg": WS_BG,
            }
        )

    # Server
    for f, nodes in sorted(site_data.get("server_ms", {}).items()):
        site_rows.append(
            {
                "file": f,
                "tipo": "Server",
                "nodi": ", ".join(nodes) if nodes else "tutti",
                "bg": SRV_BG,
            }
        )

    # Firmware
    for f, info in sorted(site_data.get("firmware", {}).items()):
        tipo = "Firmware Controller" if info.get("type") == "controller" else "Firmware I/O"
        nodi = info.get("nodes", [])
        site_rows.append(
            {
                "file": f,
                "tipo": tipo,
                "nodi": ", ".join(nodi) if nodi else "tutti i controller/IO",
                "bg": FW_BG,
            }
        )

    for r in site_rows:
        ws.cell(row=row_idx, column=1, value=r["file"]).fill = _hfill(r["bg"])
        ws.cell(row=row_idx, column=2, value=r["tipo"]).fill = _hfill(r["bg"])
        ws.cell(row=row_idx, column=3, value=r["nodi"]).fill = _hfill(r["bg"])
        for ci in range(1, 4):
            ws.cell(row=row_idx, column=ci).alignment = _wrap() if ci == 3 else _top()
        ws.row_dimensions[row_idx].height = 30
        row_idx += 1

    _set_col_widths(ws, widths)
    ws.freeze_panes = "A2"
    if site_rows:
        _add_table(ws, f"{site}_Table", f"A1:C{row_idx - 1}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def write_fermata_excel(
    patch_data: dict,
    output_path: Path | None = None,
    owners_inbox: Path | None = None,
) -> Path:
    """
    Scrive l'Excel fermata con 4 sheet.

    Args:
        patch_data: Output di kba.reporter.patch_builder.build_patch_list().
        output_path: Path di output; se None usa owners_inbox con timestamp.
        owners_inbox: Cartella di output di default.

    Returns:
        Path al file scritto.
    """
    if output_path is None:
        if owners_inbox is None:
            raise ValueError("Specificare output_path o owners_inbox")
        ts = datetime.now().strftime("%d%m%y")
        output_path = owners_inbox / f"fermata-{ts}.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # Sheet 1 — Files
    ws_files = wb.active
    ws_files.title = "Files"
    _write_files_sheet(ws_files, patch_data)

    # Sheet 2-4 — per sito
    for site in SITE_ORDER:
        site_data = patch_data.get(site, {})
        ws_site = wb.create_sheet(title=site)
        _write_site_sheet(ws_site, site, site_data)

    wb.save(output_path)
    return output_path
