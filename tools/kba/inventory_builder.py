"""
KBA Inventory Builder — Nodo × Prodotto × Versione.

Legge il merged enriched Excel e il catalogo KBA, costruisce
l'inventario di tutti i nodi con i prodotti installati e le
relative versioni note/UNKNOWN.

Output: lib/data/kba_inventory/inventory_<data>.md
"""

import openpyxl, yaml, re
from pathlib import Path
from collections import OrderedDict
from datetime import date

# ── Config ──
DELTAV_VERSIONS = {
    "Lonigo": "v14LTS",
    "Montecchio": "v15LTS",
    "Termoli": "v15LTS",
}

XLSX = Path("Library/deliverables/merged_enriched_260525.xlsx")
EXPORT_XLSX = Path("Inbox/GuardianExportFile_KnowledgeBaseArticles_260525_111240.xlsx")
RECORDS_DIR = Path("lib/data/kba_catalog/records")
OUTPUT_DIR = Path("lib/data/kba_inventory")

# Prodotti DeltaV nativi → versione = DeltaV site version
DELTAV_NATIVE_KEYWORDS = [
    "pk controller",
    "eioc",
    "controller",
    "workstation",
    "control studio",
    "proplus",
    "operator",
    "objectivity",
    "logic software",
    "function block",
    "virtual eioc",
    "veioc",
    "discovery veioc",
    "charm",
    "se6502",
    "se6504",
    "se6500",
    "ve210x",
    "ve3101",
    "ve2161",
    "ve2162",
    "inter-zone",
    "zone",
    "redundant cioc",
    "standalone pk",
    "virtualization software",
]


def is_deltav_native(prod: str) -> bool:
    p = prod.lower()
    for kw in DELTAV_NATIVE_KEYWORDS:
        if kw in p:
            return True
    return False


def load_products(slug: str) -> list[str]:
    path = RECORDS_DIR / f"{slug}.md"
    if not path.exists():
        return []
    content = path.read_text(encoding="utf-8")
    m = re.match(r"\A---\s*\n([\s\S]*?)\n---", content, re.DOTALL)
    if not m:
        return []
    fm = yaml.safe_load(m.group(1)) or {}
    return fm.get("affected_products", [])


def load_node_systems() -> dict[str, str]:
    """Costruisce mappa nodo → System Name (ID) dall'export originale."""
    wb = openpyxl.load_workbook(EXPORT_XLSX, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))]
    col = {h: i for i, h in enumerate(headers)}

    sys_col = headers.index("System Name (ID)")
    node_col = headers.index("Node Name / Node Assignment")

    node_systems: dict[str, str] = {}
    for cells in ws.iter_rows(min_row=2, values_only=False):
        vals = [c.value for c in cells]
        sys_id = str(vals[sys_col] or "").strip()
        nodes_raw = str(vals[node_col] or "").strip()
        if not sys_id or not nodes_raw:
            continue
        for node in nodes_raw.split("\n"):
            node = node.strip()
            if node and node not in node_systems:
                node_systems[node] = sys_id

    wb.close()
    return node_systems


def build(node_systems: dict[str, str] | None = None) -> list[dict]:
    """Costruisce inventario: lista di dict {node, site, system_name, product, version, firmware}."""
    if node_systems is None:
        node_systems = load_node_systems()

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))]
    col = {h: i for i, h in enumerate(headers)}

    # Raccogli { (node, site): set<prodotti> }
    inventory: dict[tuple[str, str], set[str]] = OrderedDict()

    for cells in ws.iter_rows(min_row=2, values_only=False):
        vals = [c.value for c in cells]
        kba = str(vals[col["KBA Number"]] or "").strip()
        if not kba:
            continue
        suggested = str(vals[col.get("Suggested Notes", -1)] or "").strip()
        if "{WIP}" not in suggested:
            continue

        site = str(vals[col["Site"]] or "").strip().split(" - ")[0].strip()
        nodes_raw = str(vals[col["Node Name / Node Assignment"]] or "").strip()
        if not nodes_raw:
            continue

        products = load_products(kba.lower())
        if not products:
            products = [f"KBA {kba} — vedi documento"]

        for node in nodes_raw.split("\n"):
            node = node.strip()
            if not node:
                continue
            key = (node, site)
            if key not in inventory:
                inventory[key] = set()
            for p in products:
                inventory[key].add(p)

    wb.close()

    # Appiattisci in lista ordinata
    rows = []
    for (node, site), products in inventory.items():
        site_ver = DELTAV_VERSIONS.get(site, "?")
        sys_id = node_systems.get(node, "")
        for prod in sorted(products):
            version = site_ver if is_deltav_native(prod) else "UNKNOWN"
            rows.append(
                {
                    "node": node,
                    "site": site,
                    "system_name": sys_id,
                    "product": prod,
                    "version": version,
                    "firmware": "",
                }
            )
    return rows


def write_md(rows: list[dict], output_path: Path):
    """Scrive inventario in Markdown."""
    today = date.today().isoformat()
    nodes = sorted(set(r["node"] for r in rows))
    sites = sorted(set(r["site"] for r in rows))

    # Per ogni sito, tabella compatta
    lines = [
        f"# Inventario Nodi × Prodotto",
        f"",
        f"_Generato: {today} | Fonte: `{XLSX.name}` + catalogo KBA_",
        f"_Nodi unici: {len(nodes)} | Righe prodotto: {len(rows)}_",
        f"",
    ]

    # Statistiche UNKNOWN
    unknowns = [r for r in rows if r["version"] == "UNKNOWN"]
    if unknowns:
        unknown_products = sorted(set(r["product"] for r in unknowns))
        lines.append("## Versioni UNKNOWN — Da verificare")
        lines.append("")
        lines.append("| Prodotto | Nodi coinvolti | Siti |")
        lines.append("|----------|---------------|------|")
        for prod in unknown_products:
            prod_rows = [r for r in unknowns if r["product"] == prod]
            prod_nodes = sorted(set(r["node"] for r in prod_rows))
            prod_sites = sorted(set(r["site"] for r in prod_rows))
            lines.append(f"| {prod} | {len(prod_nodes)} | {', '.join(prod_sites)} |")
        lines.append("")

    # Tabella completa
    lines.append("## Inventario completo")
    lines.append("")
    lines.append("| Nodo | Sito | System ID | Prodotto | Versione | Firmware |")
    lines.append("|------|------|-----------|----------|----------|----------|")

    prev_node = ""
    for r in rows:
        node_display = r["node"] if r["node"] != prev_node else ""
        prev_node = r["node"]
        fw = r.get("firmware", "") or ""
        sys_id = r.get("system_name", "")
        lines.append(
            f"| {node_display} | {r['site']} | {sys_id} | {r['product']} | {r['version']} | {fw} |"
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


if __name__ == "__main__":
    print("📡 Caricamento mappa nodo → System ID...")
    node_systems = load_node_systems()
    print(f"   {len(node_systems)} nodi mappati a system ID")

    print("🔧 Costruzione inventario...")
    rows = build(node_systems=node_systems)
    today = date.today().isoformat().replace("-", "")
    output = OUTPUT_DIR / f"inventory_{today}.md"
    write_md(rows, output)
    print(f"✅ Inventario scritto: {output}")
    print(f"   {len(rows)} righe, {len(set(r['node'] for r in rows))} nodi unici")
