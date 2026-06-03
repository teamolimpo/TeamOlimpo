"""Carica e classifica le righe del file Excel KBA_Merged."""

import re
from pathlib import Path

import openpyxl
from loguru import logger

from tools.kba_reporter.config import DEFER_KEYWORDS, WIP_KEYWORDS, DONE_NA_KEYWORDS


def load_excel(path: Path) -> list[dict]:
    """Carica il foglio KBA_Merged (o il primo disponibile) e restituisce le righe come dict."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["KBA_Merged"] if "KBA_Merged" in wb.sheetnames else wb.active
    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = row
            continue
        if not any(row):
            continue
        r = dict(zip(headers, row))
        rows.append(
            {
                "kba_number": str(r.get("KBA Number") or "").strip(),
                "title": str(r.get("Title") or "").strip(),
                "site": str(r.get("Site") or "").strip(),
                "node": str(r.get("Node Name / Node Assignment") or "").strip(),
                "user_notes": str(r.get("User Notes") or "").strip(),
                "risk_score": r.get("Risk Score") or "",
                "risk_level": str(r.get("Risk Level") or "").strip(),
                "emerson_cat": str(r.get("Emerson Category") or "").strip(),
                "fis_notes": str(r.get("FIS Notes") or "").strip(),
                "stefano_notes": str(r.get("Stefano's Notes") or "").strip(),
            }
        )
    wb.close()
    logger.info(f"Lette {len(rows)} righe da {path.name}")
    return rows


def _is_dominated_by_done_na(notes: str) -> bool:
    """True se il marcatore dominante (primo marcatore trovato) è DONE/NA/ACK."""
    for marker in ["{DONE}", "{NA}", "{ACK}"]:
        if marker in notes:
            return True
    return False


def _matches_keywords(text: str, keywords: list[str]) -> bool:
    low = text.lower()
    return any(kw.lower() in low for kw in keywords)


def classify_rows(rows: list[dict]) -> dict:
    """Classifica le righe in DEFER e WIP in base alle keyword nelle note utente."""
    defer, wip = [], []
    for row in rows:
        if not row["kba_number"]:
            continue
        notes = row["user_notes"]
        if _matches_keywords(notes, DEFER_KEYWORDS):
            defer.append(row)
        if _matches_keywords(notes, WIP_KEYWORDS) and not _is_dominated_by_done_na(notes):
            wip.append(row)
    logger.info(f"Classificati: {len(defer)} DEFER, {len(wip)} WIP")
    return {"defer": defer, "wip": wip}
