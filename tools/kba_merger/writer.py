"""
Scrittura dei file Excel di output per il tool kba_merger.

Produce:
  - KBA_Merged_GGMMAA_HHMMSS.xlsx  (comando merge)
  - kba_gap_GGMMAA.txt             (comando gap — report testuale)
  - kba_gap_GGMMAA.xlsx            (comando gap — report tabellare)
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from loguru import logger
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from tools.kba_merger.config import (
    COLUMN_WIDTHS,
    GAP_STATUS_COLORS,
    HEADER_BG_COLOR,
    HEADER_FG_COLOR,
    OUTPUT_DIR,
    OUTPUT_HEADERS,
    RISK_LEVEL_COLORS,
    SUGGESTED_NOTE_BG_COLOR,
)

# ---------------------------------------------------------------------------
# Utilita' stile
# ---------------------------------------------------------------------------


def _make_header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=HEADER_BG_COLOR)


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _make_header_font(color: str = HEADER_FG_COLOR) -> Font:
    return Font(bold=True, color=color)


def _make_data_font(color: str = "000000") -> Font:
    return Font(color=color)


def _wrap_alignment() -> Alignment:
    return Alignment(wrap_text=True, vertical="top")


def _top_alignment() -> Alignment:
    return Alignment(vertical="top")


# ---------------------------------------------------------------------------
# Scrittura merge Excel
# ---------------------------------------------------------------------------


def write_merge_excel(
    rows: list[dict[str, Any]],
    output_path: Path | None = None,
) -> Path:
    """
    Scrive il file Excel di merge con formattazione professionale e protezione foglio.

    Struttura:
      - 17 colonne (A→Q) con header colorato
      - Wrap text su Title, User Notes, FIS Notes, Suggested Notes, Stefano's Notes
      - Altezza righe dati: 60pt
      - Freeze riga 1
      - Tabella Excel con banding (KBA_Table)
      - Foglio protetto: solo FIS Notes, Suggested Notes e Stefano's Notes editabili
      - Colonna Risk Level colorata per livello
      - Colonna Suggested Notes con sfondo giallo chiaro

    Args:
        rows: Lista di dizionari con le colonne di output.
        output_path: Path di output; se None usa OUTPUT_DIR con timestamp.

    Returns:
        Path al file scritto.
    """
    if output_path is None:
        ts = datetime.now().strftime("%d%m%y_%H%M%S")
        output_path = OUTPUT_DIR / f"KBA_Merged_{ts}.xlsx"

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KBA_Merged"

    final_headers: list[str] = list(OUTPUT_HEADERS)
    final_widths: list[int] = list(COLUMN_WIDTHS)
    num_cols = len(final_headers)
    num_rows = len(rows)

    WRAP_HEADERS = {"Title", "User Notes", "FIS Notes", "Suggested Notes", "Stefano's Notes"}

    header_idx: dict[str, int] = {h: i + 1 for i, h in enumerate(final_headers)}
    fis_col_idx = header_idx.get("FIS Notes", 0)
    suggested_note_col_idx = header_idx.get("Suggested Notes", 0)
    stefano_notes_col_idx = header_idx.get("Stefano's Notes", 0)
    risk_level_col_idx = header_idx.get("Risk Level", 0)
    wrap_col_idxs = {header_idx[h] for h in WRAP_HEADERS if h in header_idx}

    # --- Header riga 1 ---
    header_fill = _make_header_fill()
    header_font = _make_header_font()

    for col_idx, header in enumerate(final_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = _top_alignment()
        cell.protection = Protection(locked=True)

    # --- Righe dati ---
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(final_headers, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Allineamento
            if col_idx in wrap_col_idxs:
                cell.alignment = _wrap_alignment()
            else:
                cell.alignment = _top_alignment()

            # Protezione: FIS Notes, Suggested Notes e Stefano's Notes sbloccate
            if col_idx in (fis_col_idx, suggested_note_col_idx, stefano_notes_col_idx):
                cell.protection = Protection(locked=False)
            else:
                cell.protection = Protection(locked=True)

            # Sfondo giallo chiaro per Suggested Note
            if col_idx == suggested_note_col_idx and value:
                cell.fill = _make_fill(SUGGESTED_NOTE_BG_COLOR)

            # Colore Risk Level
            if col_idx == risk_level_col_idx and value:
                level_key = str(value).lower().strip()
                colors = RISK_LEVEL_COLORS.get(level_key)
                if colors:
                    bg_hex, fg_hex = colors
                    cell.fill = _make_fill(bg_hex)
                    cell.font = _make_data_font(fg_hex)

        # Altezza riga
        ws.row_dimensions[row_idx].height = 60

    # --- Larghezze colonne ---
    for col_idx, width in enumerate(final_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Freeze riga 1 ---
    ws.freeze_panes = "A2"

    # --- Tabella Excel con banding ---
    if num_rows > 0:
        last_col_letter = get_column_letter(num_cols)
        last_row = num_rows + 1
        table_ref = f"A1:{last_col_letter}{last_row}"
        table = Table(displayName="KBA_Table", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    # --- Protezione foglio ---
    ws.protection.sheet = True
    ws.protection.enable()
    ws.protection.sort = False  # allowSort=True -> sort NON bloccato
    ws.protection.autoFilter = False  # allowFiltering=True -> filter NON bloccato

    wb.save(output_path)
    logger.info(f"Merge Excel scritto: {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# Scrittura gap report testuale
# ---------------------------------------------------------------------------


def write_gap_txt(
    gap_rows: list[dict[str, Any]],
    report_date: str,
    output_path: Path | None = None,
) -> Path:
    """
    Scrive il report testuale del gap check.

    Args:
        gap_rows: Lista di dizionari output di gap.compute_gap().
        report_date: Data del report in formato 'YYYY-MM-DD'.
        output_path: Path di output; se None usa OUTPUT_DIR con data.

    Returns:
        Path al file scritto.
    """
    if output_path is None:
        date_str = report_date.replace("-", "")[2:]  # GGMMAA da YYYYMMDD
        output_path = OUTPUT_DIR / f"kba_gap_{date_str}.txt"

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    total = len(gap_rows)
    ok_count = sum(1 for r in gap_rows if r["stato"] == "ok")
    da_rianalizzare = [r for r in gap_rows if r["stato"] == "da_rianalizzare"]
    da_analizzare = [r for r in gap_rows if r["stato"] == "da_analizzare"]
    da_convertire = [r for r in gap_rows if r["stato"] == "da_convertire"]

    lines: list[str] = [
        f"KBA Gap Report — {report_date}",
        "=" * 30,
        f"Totale KBA nel file DeltaV:    {total}",
        f"In catalogo (ok):              {ok_count}",
        f"Documento riconvertito, da rianalizzare: {len(da_rianalizzare)}",
        f"PDF convertito, manca analisi: {len(da_analizzare)}",
        f"Da convertire (PDF non trovato): {len(da_convertire)}",
        "",
    ]

    if da_convertire:
        lines.append("--- DA CONVERTIRE ---")
        for r in da_convertire:
            lines.append(r["kba_number"])
        lines.append("")

    if da_rianalizzare:
        lines.append("--- DA RIANALIZZARE ---")
        for r in da_rianalizzare:
            lines.append(r["kba_number"])
        lines.append("")

    if da_analizzare:
        lines.append("--- DA ANALIZZARE ---")
        for r in da_analizzare:
            lines.append(r["kba_number"])
        lines.append("")

    output_path.write_text("\n".join(lines), encoding="utf-8")
    logger.info(f"Gap report TXT scritto: {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# Scrittura gap report Excel
# ---------------------------------------------------------------------------


def write_gap_excel(
    gap_rows: list[dict[str, Any]],
    report_date: str,
    output_path: Path | None = None,
) -> Path:
    """
    Scrive il report Excel del gap check con colorazione per stato.

    Colonne: KBA Number, Published, Category, Disposition Status, Title,
             Occorrenze, Stato

    Args:
        gap_rows: Lista di dizionari output di gap.compute_gap().
        report_date: Data del report in formato 'YYYY-MM-DD'.
        output_path: Path di output; se None usa OUTPUT_DIR con data.

    Returns:
        Path al file scritto.
    """
    if output_path is None:
        date_str = report_date.replace("-", "")[2:]  # GGMMAA
        output_path = OUTPUT_DIR / f"kba_gap_{date_str}.xlsx"

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KBA_Gap"

    # Rileva se almeno una riga ha referenced_by valorizzato
    has_refs = any(r.get("referenced_by") for r in gap_rows)

    gap_headers = [
        "KBA Number",
        "Published",
        "Category",
        "Disposition Status",
        "Title",
        "Occorrenze",
        "Stato",
    ]
    gap_widths = [16, 12, 14, 18, 50, 12, 16]
    if has_refs:
        gap_headers.append("Referenziata da")
        gap_widths.append(30)

    # --- Header ---
    header_fill = _make_header_fill()
    header_font = _make_header_font()
    for col_idx, header in enumerate(gap_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = _top_alignment()

    # --- Dati ---
    STATO_COL_IDX = 7  # colonna G
    for row_idx, r in enumerate(gap_rows, 2):
        ws.cell(row=row_idx, column=1, value=r["kba_number"])
        ws.cell(row=row_idx, column=2, value=r["published"])
        ws.cell(row=row_idx, column=3, value=r["category"])
        ws.cell(row=row_idx, column=4, value=r["disposition_status"])

        title_cell = ws.cell(row=row_idx, column=5, value=r["title"])
        title_cell.alignment = _top_alignment()

        ws.cell(row=row_idx, column=6, value=r["occurrences"])

        stato = r["stato"]
        stato_cell = ws.cell(row=row_idx, column=STATO_COL_IDX, value=stato)
        color_hex = GAP_STATUS_COLORS.get(stato)
        if color_hex:
            stato_cell.fill = _make_fill(color_hex)

        if has_refs:
            ref_val = ", ".join(r.get("referenced_by") or [])
            ws.cell(row=row_idx, column=8, value=ref_val)

    # --- Larghezze ---
    for col_idx, width in enumerate(gap_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Freeze ---
    ws.freeze_panes = "A2"

    # --- Tabella ---
    num_rows = len(gap_rows)
    if num_rows > 0:
        last_col = "H" if has_refs else "G"
        table = Table(displayName="KBA_Gap_Table", ref=f"A1:{last_col}{num_rows + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    wb.save(output_path)
    logger.info(f"Gap report Excel scritto: {output_path}")
    return output_path
